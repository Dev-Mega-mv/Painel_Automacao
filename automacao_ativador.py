from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, InvalidSessionIdException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from collections import defaultdict
import re
import tempfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyautogui
import keyboard
import traceback
import random
import time
import os

automation_status = {
    'running': False,
    'total': 0,
    'processed': 0,
    'current_item': None,
    'logs': [],
    'start_time': None,
    'end_time': None,
    'request_stop': False,
    'relatorio_path': None,
}

STATUS_ATIVO   = ('ATIVADO', 'APROVADO', 'ENVIADO')
STATUS_INATIVO = ('', 'CANCELADO NA ADM', 'CANCELADO NA REDE')

XP_BTN_CARDSE        = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[1]/td[4]/a'
XP_BTN_CIELO         = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[2]/td[4]/a'
XP_SITUACAO_CARDSE   = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[1]/td[2]'
XP_SITUACAO_CIELO    = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[2]/td[2]'
XP_STATUS_CARDSE     = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[1]/td[3]'
XP_STATUS_CIELO      = '//*[@id="report_REDE_DE_CAPTURA"]/tbody[2]/tr/td/table/tbody/tr[2]/td[3]'
XP_VOLTAR            = '//*[@id="VOLTAR"]/span'
XP_POPUP_CIELO_OK    = '//*[@id="button-1006-btnIconEl"]'
XP_SITUACAO_CONTRATO = '/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[1]/div[2]/table/tbody/tr[1]/td[3]/span/span'

CONTRATO_STATUS_ORDEM = [
    'APROVADO', 'APROVADOS', 'ATIVADO', 'ATIVADOS', 'ENVIADO', 'ENVIADOS',
    'CANCELADO', 'CANCELADOS', 'CANCELADO NA ADM', 'CANCELADO NA REDE',
    'DEFINICAO', 'DEFINIÇÃO', 'INATIVO', 'INATIVOS', 'DESATIVADO', 'VAZIOS'
]


def _is_session_error(e):
    msg = str(e).lower()
    return (
        "invalid session id"              in msg
        or "no such window"               in msg
        or "no such session"              in msg
        or "target window already closed" in msg
        or "max retries exceeded"         in msg   # Chrome caiu (WinError 10061 / chromedriver desconectado)
        or "connection refused"           in msg   # porta do chromedriver fechada
        or isinstance(e, InvalidSessionIdException)
    )


def _normalizar_status_contrato(valor):
    valor = (valor or '').strip()
    if not valor:
        return 'DESATIVADO'
    return valor.upper()


def _ordenar_status_contrato(chave):
    try:
        return (CONTRATO_STATUS_ORDEM.index(chave), chave)
    except ValueError:
        return (len(CONTRATO_STATUS_ORDEM), chave)


def _excel_safe_value(value):
    if value is None:
        return ''
    if isinstance(value, (int, float, bool, datetime)):
        return value
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
    if isinstance(value, (list, tuple, set)):
        return ', '.join(str(_excel_safe_value(v)) for v in value)
    if isinstance(value, dict):
        return ', '.join(f"{k}: {_excel_safe_value(v)}" for k, v in value.items())
    return str(value)


def _save_workbook_atomically(wb, caminho):
    diretorio = os.path.dirname(caminho) or os.getcwd()
    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=diretorio)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, caminho)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def _validar_xlsx(caminho):
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True)
        wb.close()
        return True
    except Exception:
        return False


def _ajustar_larguras(ws):
    limites = {}
    for row in ws.iter_rows():
        for cell in row:
            valor = '' if cell.value is None else str(cell.value)
            maior_linha = max((len(parte) for parte in valor.splitlines()), default=0)
            limites[cell.column] = max(limites.get(cell.column, 0), min(maior_linha + 2, 80))
    for col_idx, largura in limites.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, largura)


def _montar_resumo_contratos(grupos):
    grupos_validos = {status: nums for status, nums in grupos.items() if nums}
    total = sum(len(nums) for nums in grupos_validos.values())
    if total == 0:
        return {
            'total': 0,
            'resumo': 'Nenhum contrato encontrado',
            'detalhes': 'Nenhum contrato encontrado',
            'status_unico': '',
        }

    ordenados = sorted(grupos_validos.items(), key=lambda item: (-len(item[1]), _ordenar_status_contrato(item[0])))
    resumo = ' | '.join(f"{status}: {len(nums)}" for status, nums in ordenados)

    if len(ordenados) == 1:
        status, nums = ordenados[0]
        if status in ('VAZIOS', 'DESATIVADO'):
            detalhes = f"Todos os {len(nums)} contratos ficaram desativados."
        else:
            detalhes = f"Todos os {len(nums)} contratos estao em {status}."
        return {
            'total': total,
            'resumo': resumo,
            'detalhes': detalhes,
            'status_unico': status,
        }

    detalhes_linhas = []
    for status, nums in ordenados:
        detalhes_linhas.append(f"{status} ({len(nums)}): {', '.join(str(num) for num in nums)}")
    return {
        'total': total,
        'resumo': resumo,
        'detalhes': '\n'.join(detalhes_linhas),
        'status_unico': '',
    }


# ─────────────────────────────────────────────────────────────
#  Geração do relatório Excel
# ─────────────────────────────────────────────────────────────

def _gerar_relatorio_xlsx(dados: list, caminho: str):
    """
    dados = [
        {'cod': '12345', 'produtos': [
            {
              'num': 1, 'nome': 'PRODUTO X',
              'cielo_sit': 'APROVADO', 'cielo_hab': 'SIM',
              'cardse_sit': 'ENVIADO',  'cardse_hab': 'SIM',
              'obs': ''
            },
        ]},
    ]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verificação"

    # Definir estilos inline para evitar problemas de hashabilidade
    def get_cor_header():
        return PatternFill("solid", fgColor="1F3864")
    def get_cor_cred():
        return PatternFill("solid", fgColor="2E75B6")
    def get_cor_sim():
        return PatternFill("solid", fgColor="C6EFCE")
    def get_cor_nao():
        return PatternFill("solid", fgColor="FFCCCC")
    def get_cor_sit_ok():
        return PatternFill("solid", fgColor="DDEBF7")
    def get_cor_vazio():
        return PatternFill("solid", fgColor="F2F2F2")
    def get_fonte_header():
        return Font(bold=True, color="FFFFFF", size=13)
    def get_fonte_cred():
        return Font(bold=True, color="FFFFFF", size=11)
    def get_fonte_normal():
        return Font(size=10)
    def get_centro():
        return Alignment(horizontal="center", vertical="center")
    def get_esquerda():
        return Alignment(horizontal="left", vertical="center")
    def get_borda():
        return Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    COLUNAS = ["#", "Produto",
               "CIELO — Situação Arquivo", "CIELO — Habilitado",
               "CARDSE — Situação Arquivo", "CARDSE — Habilitado",
               "Observação"]

    # Cabeçalho global — 7 colunas
    ws.merge_cells("A1:G1")
    ws["A1"] = _excel_safe_value(f"Relatório de Verificação — gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    ws["A1"].fill = get_cor_header()
    ws["A1"].font = get_fonte_header()
    ws["A1"].alignment = get_centro()
    ws.row_dimensions[1].height = 28

    linha = 2

    for bloco in dados:
        cod = bloco.get('cod', '—')
        produtos = bloco.get('produtos', [])

        # Linha do credenciado
        ws.merge_cells(f"A{linha}:G{linha}")
        ws[f"A{linha}"] = _excel_safe_value(f"  Credenciado: {cod}")
        ws[f"A{linha}"].fill = get_cor_cred()
        ws[f"A{linha}"].font = get_fonte_cred()
        ws[f"A{linha}"].alignment = get_esquerda()
        ws.row_dimensions[linha].height = 22
        linha += 1

        # Cabeçalho da tabela
        for col, titulo in enumerate(COLUNAS, start=1):
            c = ws.cell(row=linha, column=col, value=_excel_safe_value(titulo))
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = Font(bold=True, size=10)
            c.alignment = get_centro()
            c.border = get_borda()
        ws.row_dimensions[linha].height = 18
        linha += 1

        if not produtos:
            ws.merge_cells(f"A{linha}:G{linha}")
            ws[f"A{linha}"] = _excel_safe_value("Nenhum produto encontrado")
            ws[f"A{linha}"].alignment = get_centro()
            ws[f"A{linha}"].font = Font(italic=True, size=10, color="888888")
            linha += 1
        else:
            for p in produtos:
                cielo_sit = (p.get('cielo_sit') or '').strip().upper()
                cielo_hab = (p.get('cielo_hab') or '').strip().upper()
                cardse_sit = (p.get('cardse_sit') or '').strip().upper()
                cardse_hab = (p.get('cardse_hab') or '').strip().upper()

                cor_ch = get_cor_sim() if cielo_hab == 'SIM' else (get_cor_nao() if cielo_hab == 'NÃO' else get_cor_vazio())
                cor_sh = get_cor_sim() if cardse_hab == 'SIM' else (get_cor_nao() if cardse_hab == 'NÃO' else get_cor_vazio())
                cor_cs = get_cor_sit_ok() if cielo_sit else get_cor_vazio()
                cor_ss = get_cor_sit_ok() if cardse_sit else get_cor_vazio()

                valores = [
                    p.get('num', ''), p.get('nome', ''),
                    cielo_sit or '—', cielo_hab or '—',
                    cardse_sit or '—', cardse_hab or '—',
                    p.get('obs', '')
                ]
                cores = [None, None, cor_cs, cor_ch, cor_ss, cor_sh, None]

                for col, (val, cor) in enumerate(zip(valores, cores), start=1):
                    c = ws.cell(row=linha, column=col, value=_excel_safe_value(val))
                    c.font = get_fonte_normal()
                    c.border = get_borda()
                    c.alignment = get_centro() if col != 2 else get_esquerda()
                    if cor:
                        c.fill = cor
                ws.row_dimensions[linha].height = 16
                linha += 1

        linha += 1  # espaço entre credenciados

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 28

    _save_workbook_atomically(wb, caminho)


def _gerar_relatorio_coleta_xlsx(dados: list, caminho: str, meta: dict):
    wb = openpyxl.Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_dados = wb.create_sheet("Coleta")

    # Definir estilos inline para evitar problemas de hashabilidade
    def get_cor_titulo():
        return PatternFill("solid", fgColor="0F3D5E")
    def get_cor_header():
        return PatternFill("solid", fgColor="0F766E")
    def get_cor_sub():
        return PatternFill("solid", fgColor="D9F99D")
    def get_cor_info():
        return PatternFill("solid", fgColor="F0FDFA")
    def get_cor_erro():
        return PatternFill("solid", fgColor="FEF2F2")
    def get_cor_ok():
        return PatternFill("solid", fgColor="ECFDF5")
    def get_fonte_branca():
        return Font(bold=True, color="FFFFFF", size=12)
    def get_fonte_header():
        return Font(bold=True, color="FFFFFF", size=10)
    def get_fonte_bold():
        return Font(bold=True, size=10)
    def get_centro():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def get_esquerda():
        return Alignment(horizontal="left", vertical="top", wrap_text=True)
    def get_borda():
        return Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    resumo_campos = [
        ("Gerado em", meta.get('gerado_em') or datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ("Origem alvo", meta.get('origem_transacao_alvo', '')),
        ("Tipo de pesquisa", meta.get('tipo_pesquisa', '')),
        ("Processados", meta.get('processados', 0)),
        ("Total previsto", meta.get('total', 0)),
        ("Ultimo credenciado", meta.get('ultimo_credenciado', '')),
        ("Ultimo ponto de parada", meta.get('ultima_parada', '')),
    ]

    ws_resumo.merge_cells("A1:B1")
    ws_resumo["A1"] = _excel_safe_value("Resumo da Coleta de Codigo Estabelecimento")
    ws_resumo["A1"].fill = get_cor_titulo()
    ws_resumo["A1"].font = get_fonte_branca()
    ws_resumo["A1"].alignment = get_centro()
    ws_resumo.row_dimensions[1].height = 26

    for idx, (campo, valor) in enumerate(resumo_campos, start=3):
        c1 = ws_resumo.cell(row=idx, column=1, value=_excel_safe_value(campo))
        c2 = ws_resumo.cell(row=idx, column=2, value=_excel_safe_value(valor))
        for cell in (c1, c2):
            cell.border = get_borda()
            cell.alignment = get_esquerda()
        c1.fill = get_cor_sub()
        c1.font = get_fonte_bold()
        c2.fill = get_cor_info()

    if meta.get('erros'):
        linha = len(resumo_campos) + 5
        ws_resumo.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
        ws_resumo.cell(row=linha, column=1, value=_excel_safe_value("Ocorrencias registradas")).fill = get_cor_titulo()
        ws_resumo.cell(row=linha, column=1).font = get_fonte_branca()
        ws_resumo.cell(row=linha, column=1).alignment = get_esquerda()
        linha += 1
        for erro in meta['erros']:
            ws_resumo.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
            cel = ws_resumo.cell(row=linha, column=1, value=_excel_safe_value(erro))
            cel.fill = get_cor_erro()
            cel.border = get_borda()
            cel.alignment = get_esquerda()
            linha += 1

    colunas = [
        "Codigo Infox/CNPJ", "Origem de Transacao", "Codigo Estabelecimento",
        "Data Cadastro", "Data Alteracao", "Situacao",
        "Contratos Totais", "Resumo Contratos", "Detalhes Contratos", "Observacao"
    ]
    for col, titulo in enumerate(colunas, start=1):
        cel = ws_dados.cell(row=1, column=col, value=_excel_safe_value(titulo))
        cel.fill = get_cor_header()
        cel.font = get_fonte_header()
        cel.alignment = get_centro()
        cel.border = get_borda()

    for linha, bloco in enumerate(dados, start=2):
        codigo_est = bloco.get('codigo_estabelecimento', {})
        contratos = bloco.get('contratos', {})
        valores = [
            bloco.get('cod', ''),
            codigo_est.get('origem_transacao', ''),
            codigo_est.get('codigo_estabelecimento', ''),
            codigo_est.get('data_cadastro', ''),
            codigo_est.get('data_alteracao', ''),
            codigo_est.get('situacao', ''),
            contratos.get('total', 0),
            contratos.get('resumo', ''),
            contratos.get('detalhes', ''),
            bloco.get('observacao', ''),
        ]
        tem_erro = bool(bloco.get('observacao'))
        for col, valor in enumerate(valores, start=1):
            cel = ws_dados.cell(row=linha, column=col, value=_excel_safe_value(valor))
            cel.border = get_borda()
            cel.alignment = get_esquerda() if col != 7 else get_centro()
            if linha % 2 == 0:
                cel.fill = get_cor_info()
            if tem_erro and col == 10:
                cel.fill = get_cor_erro()
            elif col in (7, 8) and not tem_erro:
                cel.fill = get_cor_ok()
        ws_dados.row_dimensions[linha].height = 42 if '\n' in str(contratos.get('detalhes', '')) else 24

    _ajustar_larguras(ws_resumo)
    _ajustar_larguras(ws_dados)
    ws_dados.freeze_panes = "A2"
    _save_workbook_atomically(wb, caminho)


# ─────────────────────────────────────────────────────────────
#  Classe principal
# ─────────────────────────────────────────────────────────────

class WebAppAtivador:
    def __init__(self, dados_login, lista_credenciados, tipo_pesquisa='COD',
                 tecnologia_padrao=1, acao_cielo='ATIVAR', acao_cardse='ATIVAR',
                 cardse_statuses=None,
                 verificar_origem=False,
                 modo_verificacao=False,
                 produto_inicio=1,
                 modo_coleta_estabelecimento=False,
                 origem_transacao_alvo='CARDSE'):
        self.dados_login        = dados_login
        self.lista_credenciados = lista_credenciados
        self.tipo_pesquisa      = tipo_pesquisa.upper()
        self.tecnologia_padrao  = tecnologia_padrao
        self.acao_cielo_padrao  = acao_cielo.upper()
        self.acao_cardse_padrao = acao_cardse.upper()
        self.cardse_statuses    = {str(v).strip().upper() for v in (cardse_statuses or []) if str(v).strip()}
        self.verificar_origem   = verificar_origem
        self.modo_verificacao   = modo_verificacao
        self.modo_coleta_estabelecimento = modo_coleta_estabelecimento
        self.origem_transacao_alvo = (origem_transacao_alvo or '').strip()
        self.produto_inicio     = max(1, int(produto_inicio))
        self.driver  = None
        self.wait    = None
        self.actions = None
        self._dados_relatorio   = []   # acumula para o xlsx
        self._relatorio_path    = None
        self._meta_relatorio = {
            'gerado_em': '',
            'origem_transacao_alvo': self.origem_transacao_alvo,
            'tipo_pesquisa': self.tipo_pesquisa,
            'processados': 0,
            'total': len(self.lista_credenciados),
            'ultimo_credenciado': '',
            'ultima_parada': '',
            'erros': [],
        }

    # ── Logging ───────────────────────────────────────────────

    def add_log(self, message, level="INFO"):
        timestamp = datetime.now().strftime('%H:%M:%S')
        prefixes  = {"ERROR": "ERRO", "WARNING": "AVISO", "SUCCESS": "SUCESSO"}
        prefix    = prefixes.get(level, "")
        log_entry = f"[{timestamp}] {f'{prefix}: ' if prefix else ''}{message}"
        automation_status['logs'].append(log_entry)
        print(log_entry)
        try:
            with open('ativador_log.txt', 'a', encoding='utf-8') as f:
                f.write(log_entry + '\n')
        except: pass

    def handle_exception(self, e, context=""):
        error_str = str(e)
        if   "stale element reference"   in error_str: msg = "Elemento desatualizado"
        elif "no such element"           in error_str: msg = "Elemento não encontrado"
        elif "timeout"          in error_str.lower():  msg = "Tempo de espera excedido"
        elif "element click intercepted" in error_str: msg = "Clique interceptado"
        elif "element not interactable"  in error_str: msg = "Elemento não interagível"
        elif _is_session_error(e):                     msg = "Sessão do navegador perdida (janela fechada?)"
        else: msg = error_str.split('\n')[0][:120]
        self.add_log(f"{context}: {msg}", level="ERROR")
        try:
            with open('debug_ativador.txt', 'a') as f:
                f.write(f"\n--- {datetime.now()} ---\n{context}\n{error_str}\n")
                traceback.print_exc(file=f)
        except: pass

    # ── Driver ────────────────────────────────────────────────

    def _iniciar_driver(self):
        self.add_log("Iniciando navegador...")
        try:
            user_data_dir = os.path.join(os.getcwd(), "selenium-profile")
            os.makedirs(user_data_dir, exist_ok=True)
            opts = Options()
            opts.add_argument(f"--user-data-dir={user_data_dir}")
            opts.add_argument("--start-maximized")
            opts.add_argument("--disable-extensions")
            opts.add_argument("--disable-popup-blocking")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.binary_location = r"C:\BACKUP_EMANUEL\Program Files\Google\Chrome\Application\chrome.exe"
            service = ChromeService(ChromeDriverManager().install())
            self.driver  = webdriver.Chrome(service=service, options=opts)
            self.driver.get("about:blank")
            self.wait    = WebDriverWait(self.driver, 15)
            self.actions = ActionChains(self.driver)
            self.driver.implicitly_wait(8)
            self._implicit_wait = 8
            time.sleep(2)
            self._navegar_pyautogui("https://online.fwcard.com.br/fwcard/f")
            time.sleep(3)
            self.add_log("Navegador iniciado.")
        except Exception as e:
            self.handle_exception(e, "Erro ao iniciar navegador")

    def _navegar_pyautogui(self, url):
        try:
            pyautogui.click(100, 100); time.sleep(.5)
            pyautogui.hotkey('ctrl', 'l'); time.sleep(.5)
            pyautogui.hotkey('ctrl', 'a'); time.sleep(.3)
            pyautogui.press('delete'); time.sleep(.3)
            pyautogui.write(url, interval=0.05)
            keyboard.press_and_release('shift+/')
            keyboard.write('p=380:100:')
            pyautogui.press('enter')
            time.sleep(3)
        except Exception as e:
            self.handle_exception(e, "Erro PyAutoGUI")

    def _realizar_login(self):
        self.add_log("Realizando login...")
        try:
            u = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="P100_USERNAME"]')))
            self._type(self.dados_login['login_infox'], u)
            p = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="P100_PASSWORD"]')))
            self._type(self.dados_login['senha_infox'], p)
            btn = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="LOGIN"]/span')))
            try:
                cap = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="uLogin"]/div[5]/div/div/div[1]')))
                self._click(cap)
            except: pass
            self._click(btn)
            self.add_log("Login realizado.")
            time.sleep(2)
        except Exception as e:
            self.handle_exception(e, "Erro no login")

    def _type(self, text, el):
        el.clear(); time.sleep(.2)
        for c in text:
            el.send_keys(c)
            time.sleep(random.uniform(.03, .1))

    def _click(self, el):
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(.2)
        ActionChains(self.driver).move_to_element(el).pause(.1).perform()
        el.click(); time.sleep(.2)

    def _click_xpath_robusto(self, xpath, descricao, timeout=10):
        ultimo_erro = None
        fim = time.time() + timeout
        while time.time() < fim:
            try:
                elementos = self.driver.find_elements(By.XPATH, xpath)
                for el in elementos:
                    if not el.is_displayed():
                        continue
                    try:
                        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    except Exception:
                        pass
                    try:
                        ActionChains(self.driver).move_to_element(el).pause(.2).click(el).perform()
                        time.sleep(.2)
                        return el
                    except Exception as e:
                        ultimo_erro = e
                    try:
                        self.driver.execute_script(
                            "arguments[0].dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));"
                            "arguments[0].dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));"
                            "arguments[0].dispatchEvent(new MouseEvent('click', {bubbles:true}));",
                            el
                        )
                        time.sleep(.2)
                        return el
                    except Exception as e:
                        ultimo_erro = e
                    try:
                        self.driver.execute_script("arguments[0].click();", el)
                        time.sleep(.2)
                        return el
                    except Exception as e:
                        ultimo_erro = e
            except Exception as e:
                ultimo_erro = e
            time.sleep(.3)
        if ultimo_erro:
            raise ultimo_erro
        raise TimeoutException(f"Clique falhou: {descricao}")

    def _aguardar_elemento_visivel(self, xpath, descricao, timeout=8):
        fim = time.time() + timeout
        ultimo_erro = None
        while time.time() < fim:
            try:
                elementos = self.driver.find_elements(By.XPATH, xpath)
                for el in elementos:
                    if el.is_displayed() and el.size.get('width', 0) > 0 and el.size.get('height', 0) > 0:
                        return el
            except Exception as e:
                ultimo_erro = e
            time.sleep(0.2)
        if ultimo_erro:
            raise ultimo_erro
        raise TimeoutException(f"Elemento nao ficou visivel: {descricao}")

    def _texto_elemento(self, el):
        try:
            links = el.find_elements(By.XPATH, './/a')
            if links and links[0].text.strip():
                return links[0].text.strip()
        except Exception:
            pass
        return el.text.strip()

    def _clicar_primeiro_disponivel(self, xpaths, descricao, timeout=8):
        fim = time.time() + timeout
        ultimo_erro = None
        while time.time() < fim:
            for xp in xpaths:
                try:
                    elementos = self.driver.find_elements(By.XPATH, xp)
                    for el in elementos:
                        if el.is_displayed() and el.is_enabled():
                            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                            self.driver.execute_script("arguments[0].click();", el)
                            time.sleep(1)
                            return
                except Exception as e:
                    ultimo_erro = e
            time.sleep(0.3)
        if ultimo_erro:
            raise ultimo_erro
        raise TimeoutException(f"Elemento nao encontrado: {descricao}")

    def _achar_item_menu_visivel(self, textos, timeout=8):
        if isinstance(textos, str):
            textos = [textos]
        fim = time.time() + timeout
        while time.time() < fim:
            for texto in textos:
                xpath = (
                    "//*[(self::span or self::a or self::div)"
                    f" and contains(normalize-space(.), \"{texto}\")]"
                )
                try:
                    elementos = self.driver.find_elements(By.XPATH, xpath)
                    for el in elementos:
                        if el.is_displayed() and el.size.get('height', 0) > 0 and el.size.get('width', 0) > 0:
                            return el
                except Exception:
                    pass
            time.sleep(0.2)
        raise TimeoutException(f"Item de menu nao encontrado: {textos}")

    def _hover_menu(self, el):
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        try:
            ActionChains(self.driver).move_to_element(el).pause(0.6).perform()
        except Exception:
            pass
        try:
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new MouseEvent('mouseover', {bubbles:true}));"
                "arguments[0].dispatchEvent(new MouseEvent('mouseenter', {bubbles:true}));",
                el
            )
        except Exception:
            pass
        time.sleep(0.8)

    def _abrir_menu_codigo_estabelecimento(self):
        self.add_log("[COLETA] Abrindo menu Opcoes...")
        self._click_xpath_robusto('//*[@id="CREDENCIADO"]/span', "Opcoes", timeout=10)
        self._aguardar_elemento_visivel(
            '//*[@id="CREDENCIADO_MENU_menu_1i"] | //*[@id="CREDENCIADO_MENU_menu_1i"]/a | //*[@id="CREDENCIADO_MENU_menu_1i"]/span | //*[@id="CREDENCIADO_MENU_menu_1"]',
            "submenu Cadastro apos clicar em Opcoes",
            timeout=5
        )
        self.add_log("[COLETA] Menu Opcoes aberto com sucesso.")
        time.sleep(0.2)

        self.add_log("[COLETA] Clicando em Cadastro...")
        self._click_xpath_robusto('//*[@id="CREDENCIADO_MENU_menu_1i"] | //*[@id="CREDENCIADO_MENU_menu_1"]', "Cadastro", timeout=10)
        self._aguardar_elemento_visivel(
            '//*[@id="CREDENCIADO_MENU_menu_1_2i"] | //*[@id="CREDENCIADO_MENU_menu_1_2"]',
            "submenu Codigo Estabelecimento apos clicar em Cadastro",
            timeout=5
        )
        self.add_log("[COLETA] Submenu Cadastro aberto com sucesso.")
        time.sleep(0.2)

        self.add_log("[COLETA] Clicando em Codigo Estabelecimento...")
        self._click_xpath_robusto('//*[@id="CREDENCIADO_MENU_menu_1_2i"] | //*[@id="CREDENCIADO_MENU_menu_1_2"]', "Codigo Estabelecimento", timeout=10)
        time.sleep(0.2)

    # ── Fluxo principal ───────────────────────────────────────

    def iniciar(self):
        try:
            if self.modo_coleta_estabelecimento:
                modo = "COLETA CODIGO ESTABELECIMENTO"
            else:
                modo = "VERIFICAÇÃO" if self.modo_verificacao else "ATIVAÇÃO"
            self.add_log(f"Automação iniciada — Modo: {modo}")
            if self.modo_coleta_estabelecimento:
                self.add_log(f"Origem da transação alvo: {self.origem_transacao_alvo or 'NÃO INFORMADA'}")
            elif self.produto_inicio > 1:
                self.add_log(f"Início a partir do produto #{self.produto_inicio}")
            automation_status['running']        = True
            automation_status['start_time']     = datetime.now()
            automation_status['total']          = len(self.lista_credenciados)
            automation_status['processed']      = 0
            automation_status['relatorio_path'] = None
            if self.modo_coleta_estabelecimento:
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                self._relatorio_path = os.path.join(os.getcwd(), f"coleta_codigo_estabelecimento_{ts}.xlsx")
                automation_status['relatorio_path'] = self._relatorio_path
            self._iniciar_driver()
            self._realizar_login()
            self.add_log("Aguardando carregamento da página inicial...")
            time.sleep(5)
            if self.modo_coleta_estabelecimento:
                self._processar_todos_v2()
            else:
                self._processar_todos()
        except Exception as e:
            self.handle_exception(e, "Erro crítico")
        finally:
            try:
                self._finalizar()
            except Exception as e:
                self.handle_exception(e, "Erro ao finalizar automação")
            automation_status['running']  = False
            automation_status['end_time'] = datetime.now()

    def _processar_todos(self):
        for idx, item in enumerate(self.lista_credenciados):
            if automation_status['request_stop']:
                self.add_log("Interrupção solicitada.")
                break

            if isinstance(item, dict):
                cod         = item['cod']
                tecnologia  = item.get('tecnologia',  self.tecnologia_padrao)
                acao_cielo  = item.get('acao_cielo',  self.acao_cielo_padrao).upper()
                acao_cardse = item.get('acao_cardse', self.acao_cardse_padrao).upper()
            else:
                cod         = str(item).strip()
                tecnologia  = self.tecnologia_padrao
                acao_cielo  = self.acao_cielo_padrao
                acao_cardse = self.acao_cardse_padrao

            tec_nome = {1:'CIELO', 2:'CARDSE', 3:'AMBOS'}.get(tecnologia, '?')
            automation_status['current_item'] = cod
            self.add_log(f"{'─'*40}")
            self.add_log(f"Credenciado [{idx+1}/{len(self.lista_credenciados)}]: {cod} | Tec: {tec_nome}")

            try:
                self._navegar_para_pesquisa()
                self._pesquisar_credenciado(cod)
                self._acessar_contratos()

                if self.modo_verificacao:
                    produtos_lidos = self._verificar_produtos(cod)
                    self._dados_relatorio.append({'cod': cod, 'produtos': produtos_lidos})
                else:
                    self._processar_produtos(cod, tecnologia, acao_cielo, acao_cardse)

                self._voltar_tela_credenciado()
                automation_status['processed'] += 1
                self.add_log(f"Credenciado {cod} finalizado!", level="SUCCESS")
            except Exception as e:
                if _is_session_error(e):
                    self.add_log("Navegador fechado inesperadamente — encerrando automação.", level="ERROR")
                    return
                self.handle_exception(e, f"Erro ao processar {cod}")
                time.sleep(3)

    # ── Navegação ─────────────────────────────────────────────

    def _processar_todos_v2(self):
        for idx, item in enumerate(self.lista_credenciados):
            if automation_status['request_stop']:
                self.add_log("Interrupção solicitada.")
                break

            cod = item['cod'] if isinstance(item, dict) else str(item).strip()
            automation_status['current_item'] = cod
            self._meta_relatorio['ultimo_credenciado'] = cod
            self._meta_relatorio['ultima_parada'] = f"Em processamento: {cod}"
            self.add_log(f"{'─' * 40}")
            self.add_log(f"Credenciado [{idx+1}/{len(self.lista_credenciados)}]: {cod} | Coleta: {self.origem_transacao_alvo}")

            try:
                self._navegar_para_pesquisa()
                self._pesquisar_credenciado(cod)
                self._processar_coleta_estabelecimento(cod)
                automation_status['processed'] += 1
                self._meta_relatorio['processados'] = automation_status['processed']
                self._meta_relatorio['ultima_parada'] = f"Concluido: {cod}"
                self._salvar_relatorio_coleta(parcial=True, silencioso=True)
                self.add_log(f"Credenciado {cod} finalizado!", level="SUCCESS")
            except Exception as e:
                if _is_session_error(e):
                    self._meta_relatorio['ultima_parada'] = f"Sessao encerrada em: {cod}"
                    self.add_log("Navegador fechado inesperadamente — encerrando automação.", level="ERROR")
                    return
                msg = str(e).split('\n')[0][:180]
                self._registrar_erro_coleta(cod, msg)
                automation_status['processed'] += 1
                self._meta_relatorio['processados'] = automation_status['processed']
                self._meta_relatorio['ultima_parada'] = f"Erro em: {cod}"
                self._salvar_relatorio_coleta(parcial=True, silencioso=True)
                self.handle_exception(e, f"Erro ao processar {cod}")
                time.sleep(3)

    def _navegar_para_pesquisa(self):
        self.add_log("Navegando para pesquisa...")
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t_MenuNav_11i"]'))).click();       time.sleep(1)
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t_MenuNav_11_1i"]'))).click();     time.sleep(1)
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t_MenuNav_11_1_2i"]'))).click();   time.sleep(1)
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t_MenuNav_11_1_2_1i"]'))).click(); time.sleep(1)
            self.add_log("Navegação concluída.")
        except Exception as e:
            self.handle_exception(e, "Erro na navegação"); raise

    def _pesquisar_credenciado(self, codigo):
        self.add_log(f"Pesquisando: {codigo} ({self.tipo_pesquisa})")
        try:
            xp = ('/html/body/div[1]/div/form/div/div[1]/section/div[2]/section[1]/div[2]/table/tbody/tr[1]/td[1]/span/div/input'
                  if self.tipo_pesquisa == "COD" else
                  '/html/body/div[1]/div/form/div/div[1]/section/div[2]/section[1]/div[2]/table/tbody/tr[1]/td[2]/span/input')
            campo = self.wait.until(EC.presence_of_element_located((By.XPATH, xp)))
            campo.clear(); campo.send_keys(str(codigo))
            self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PESQUISAR"]/span'))).click()
            time.sleep(1)
        except Exception as e:
            self.handle_exception(e, f"Erro ao pesquisar {codigo}"); raise

    def _acessar_contratos(self):
        try:
            self._click_xpath_robusto('//*[@id="VISUALIZAR_CONTRATOS"] | //*[@id="VISUALIZAR_CONTRATOS"]/span', "Visualizar Contratos", timeout=10)
            time.sleep(1)
            self._garantir_ordem_crescente()
        except Exception as e:
            self.handle_exception(e, "Erro ao acessar contratos"); raise

    def _garantir_ordem_crescente(self):
        """Garante ordem crescente verificando se o primeiro número da lista é 1.
        Se não for, clica no header Número e verifica novamente — repete até confirmar."""
        XP_HEADER   = '//*[@id="NRO_CONTRATO"]/div/span[1]/a'
        XP_PRIMEIRO = '/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[2]/div[2]/table/tbody[2]/tr/td/table/tbody/tr[1]/td[2]/a'
        MAX_TENTATIVAS = 4
        try:
            WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.XPATH, XP_HEADER))
            )
            for tentativa in range(1, MAX_TENTATIVAS + 1):
                els = self.driver.find_elements(By.XPATH, XP_PRIMEIRO)
                if not els:
                    self.add_log("Aviso: não foi possível ler o primeiro número da lista.")
                    break
                try:
                    primeiro_num = int(els[0].text.strip())
                except ValueError:
                    primeiro_num = -1

                if primeiro_num == 1:
                    self.add_log(f"Ordenação crescente confirmada (primeiro=#1).")
                    break

                self.add_log(f"Tentativa {tentativa}: primeiro número é #{primeiro_num}, não é #1 → clicando em Número...")
                header = self.driver.find_element(By.XPATH, XP_HEADER)
                self.driver.execute_script("arguments[0].click();", header)
                time.sleep(1.5)
            else:
                self.add_log("Aviso: não foi possível garantir ordenação crescente após tentativas.")
        except Exception as e:
            self.handle_exception(e, "Aviso: erro ao garantir ordenação crescente")

    def _voltar_tela_credenciado(self):
        try:
            btn = self.wait.until(EC.element_to_be_clickable((By.XPATH,
                '/html/body/div[2]/div/form/div/div[1]/section/div[1]/span/button/span')))
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(1)
        except Exception as e:
            self.handle_exception(e, "Erro ao voltar")

    def _processar_coleta_estabelecimento(self, cod):
        dados_codigo = self._coletar_codigo_estabelecimento(cod)
        situacao_codigo = (dados_codigo.get('situacao') or '').strip().upper()
        if situacao_codigo in ('INATIVO', 'INATIVOS'):
            self.add_log(f"[COLETA] Código Estabelecimento inativo para {cod}; pulando coleta de contratos.", level="INFO")
            resumo_contratos = {
                'total': 0,
                'grupos': {},
                'resumo': 'Contratos não coletados - código inativo',
                'detalhes': 'Contratos não foram acessados porque o código estabelecimento está inativo.',
                'status_unico': 'INATIVO',
                'observacao': '',
            }
            observacoes = [dados_codigo.get('observacao', ''), resumo_contratos.get('observacao', '')]
            observacao = ' | '.join(parte for parte in observacoes if parte)
            self._dados_relatorio.append({
                'cod': cod,
                'codigo_estabelecimento': dados_codigo,
                'contratos': resumo_contratos,
                'observacao': observacao,
            })
            return

        self._acessar_contratos()
        resumo_contratos = self._coletar_resumo_contratos(cod)
        self._voltar_tela_credenciado()
        observacoes = [dados_codigo.get('observacao', ''), resumo_contratos.get('observacao', '')]
        observacao = ' | '.join(parte for parte in observacoes if parte)
        self._dados_relatorio.append({
            'cod': cod,
            'codigo_estabelecimento': dados_codigo,
            'contratos': resumo_contratos,
            'observacao': observacao,
        })

    def _coletar_codigo_estabelecimento(self, cod):
        self._abrir_menu_codigo_estabelecimento()
        self.add_log("[COLETA] Aguardando tela de Codigo Estabelecimento...")
        self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="report_CODIGO_ESTABELECIMENTO"]')))
        self.add_log(f"[COLETA] Buscando origem '{self.origem_transacao_alvo}' na tela de Codigo Estabelecimento...")

        alvo = (self.origem_transacao_alvo or '').strip().upper()
        encontrado = {
            'origem_transacao': '',
            'codigo_estabelecimento': '',
            'data_cadastro': '',
            'data_alteracao': '',
            'situacao': '',
            'observacao': '',
        }

        try:
            self.driver.implicitly_wait(0)
            linhas = self.driver.find_elements(By.XPATH, '//*[@id="report_CODIGO_ESTABELECIMENTO"]//tr[td]')
            self.add_log(f"[COLETA] {len(linhas)} linhas candidatas encontradas na tabela.")

            for linha in linhas:
                colunas = linha.find_elements(By.XPATH, './td')
                if len(colunas) < 6:
                    continue
                textos = [self._texto_elemento(col).strip() for col in colunas]
                origem_idx = -1
                for idx, texto in enumerate(textos):
                    texto_norm = texto.upper()
                    if texto_norm == alvo or (alvo and alvo in texto_norm):
                        origem_idx = idx
                        break
                if origem_idx < 0:
                    continue

                if origem_idx + 4 >= len(colunas):
                    self.add_log(f"[COLETA] Linha ignorada por estrutura inesperada: {textos}", level="WARNING")
                    continue

                origem = textos[origem_idx]
                origem_norm = origem.upper()
                if origem_norm != alvo and alvo not in origem_norm:
                    continue
                encontrado = {
                    'origem_transacao': origem,
                    'codigo_estabelecimento': textos[origem_idx + 1],
                    'data_cadastro': textos[origem_idx + 2],
                    'data_alteracao': textos[origem_idx + 3],
                    'situacao': textos[origem_idx + 4],
                    'observacao': '',
                }
                self.add_log(f"[COLETA] Origem encontrada para {cod}: {origem} | Codigo Estabelecimento: {encontrado['codigo_estabelecimento']}")
                break
        finally:
            self.driver.implicitly_wait(self._implicit_wait)

        if not encontrado['origem_transacao']:
            encontrado['observacao'] = f"Origem '{self.origem_transacao_alvo}' não encontrada."
            self.add_log(f"[COLETA] Origem '{self.origem_transacao_alvo}' não encontrada para {cod}.", level="WARNING")

        self._click_xpath_robusto(
            '//*[@id="P2118_VOLTAR"] | //*[@id="P2118_VOLTAR"]/span | //span[normalize-space(.)="Voltar"] | //button[normalize-space(.)="Voltar"]',
            "Voltar da tela de Codigo Estabelecimento",
            timeout=10
        )
        return encontrado

    def _coletar_resumo_contratos(self, cod):
        linha_idx = 1
        grupos = defaultdict(list)
        self.add_log(f"[COLETA] Lendo contratos do credenciado {cod}...")

        while True:
            if automation_status['request_stop']:
                break
            try:
                base = f'/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[2]/div[2]/table/tbody[2]/tr/td/table/tbody/tr[{linha_idx}]'
                num_el = self.driver.find_elements(By.XPATH, base + '/td[2]/a')
                if not num_el or not num_el[0].text.strip():
                    break

                numero = num_el[0].text.strip()
                try:
                    numero_int = int(numero)
                except ValueError:
                    numero_int = numero

                num_el[0].click()
                self._aguardar_detalhe_produto(timeout=5)
                situacao = _normalizar_status_contrato(self._get_status(XP_SITUACAO_CARDSE))
                grupos[situacao].append(numero_int)
                self.add_log(f"  Contrato {numero_int}: {situacao}")
                self._voltar_para_lista_produtos()
                linha_idx += 1
            except (NoSuchElementException, TimeoutException):
                break
            except Exception as e:
                if _is_session_error(e):
                    raise
                self.handle_exception(e, f"Erro coleta contrato linha #{linha_idx}")
                time.sleep(2)
                linha_idx += 1

        resumo = _montar_resumo_contratos(grupos)
        self.add_log(f"[COLETA] {resumo['total']} contratos lidos para {cod}.")
        return {
            'total': resumo['total'],
            'grupos': {status: grupos[status] for status in sorted(grupos, key=_ordenar_status_contrato)},
            'resumo': resumo['resumo'],
            'detalhes': resumo['detalhes'],
            'status_unico': resumo['status_unico'],
            'observacao': '',
        }

    def _registrar_erro_coleta(self, cod, mensagem):
        texto = f"{cod}: {mensagem}"
        self._meta_relatorio['erros'].append(texto)
        self._dados_relatorio.append({
            'cod': cod,
            'codigo_estabelecimento': {
                'origem_transacao': self.origem_transacao_alvo,
                'codigo_estabelecimento': '',
                'data_cadastro': '',
                'data_alteracao': '',
                'situacao': '',
                'observacao': '',
            },
            'contratos': {
                'total': 0,
                'grupos': {},
                'resumo': '',
                'detalhes': '',
                'status_unico': '',
                'observacao': '',
            },
            'observacao': f"Erro ao processar credenciado: {mensagem}",
        })

    def _salvar_relatorio_coleta(self, parcial=False, silencioso=False):
        if not self._relatorio_path:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self._relatorio_path = os.path.join(os.getcwd(), f"coleta_codigo_estabelecimento_{ts}.xlsx")

        self._meta_relatorio['gerado_em'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        self._meta_relatorio['processados'] = automation_status['processed']
        self._meta_relatorio['total'] = automation_status['total']
        _gerar_relatorio_coleta_xlsx(self._dados_relatorio, self._relatorio_path, self._meta_relatorio)
        if not parcial and not _validar_xlsx(self._relatorio_path):
            raise RuntimeError(f"Falha na validação do arquivo de relatório: {self._relatorio_path}")
        automation_status['relatorio_path'] = self._relatorio_path
        if parcial and not silencioso:
            self.add_log(f"Planilha parcial salva: {os.path.basename(self._relatorio_path)}")

    def _get_status(self, xp):
        try:
            return self.driver.find_elements(By.XPATH, xp)[0].text.strip()
        except:
            return ''

    def _cardse_situacao_permitida(self, situacao):
        situacao = (situacao or '').strip().upper()

        if not self.cardse_statuses:
            return False

        for status_permitido in self.cardse_statuses:
            status_permitido = status_permitido.strip().upper()

            if status_permitido in situacao:
                return True

        return False

    def _aguardar_status(self, xp, esperado, timeout=4):
        esperado = esperado.upper()
        fim = time.time() + timeout
        ultimo = ''
        while time.time() < fim:
            ultimo = (self._get_status(xp) or '').upper()
            if esperado in ('NAO', 'NÃO') and ultimo and ultimo != 'SIM':
                return True
            if ultimo == esperado:
                return True
            time.sleep(0.2)
        return False

    def _clicar_rede_captura(self, xp_botao, xp_status, esperado, descricao, timeout=4):
        self.driver.find_element(By.XPATH, xp_botao).click()
        return self._aguardar_status(xp_status, esperado, timeout=timeout)

    def _voltar_para_lista_produtos(self, timeout=8):
        self.driver.find_element(By.XPATH, XP_VOLTAR).click()
        WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[2]/div[2]/table/tbody[2]/tr/td/table/tbody/tr[1]/td[2]/a'))
        )

    def _aguardar_detalhe_produto(self, timeout=6):
        WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, XP_STATUS_CARDSE))
        )

    def _aprovar_se_necessario(self):
        try:
            sit_el = WebDriverWait(self.driver, 3).until(
                EC.presence_of_element_located((By.XPATH,
                    '/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[1]/div[2]/table/tbody/tr[1]/td[3]/span/span'
                ))
            )
            if sit_el.text.strip() == "DEFINIÇÃO":
                self.add_log("Aprovando produto em DEFINIÇÃO...")
                self._click(self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="CONTRATOS"]/span'))))
                self._click(self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="CONTRATOS_MENU_menu_1i"]'))))
                try:
                    obs = self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="P2111_OBS1"]')))
                    obs.clear()
                    obs.send_keys('Ativação realizada por bot')
                except Exception:
                    pass
                self._click(self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="OK"]/span'))))
                try:
                    WebDriverWait(self.driver, 3).until(EC.element_to_be_clickable((By.XPATH, XP_VOLTAR)))
                except TimeoutException:
                    pass
        except (TimeoutException, NoSuchElementException):
            pass
        except Exception as e:
            self.handle_exception(e, "Erro ao aprovar produto")

    # ── Modo VERIFICAÇÃO ──────────────────────────────────────

    def _verificar_produtos(self, cod) -> list:
        linha_idx = 1   # índice da linha na tabela HTML (sempre sobe de 1 em 1)
        resultado = []
        self.add_log(f"[VERIFICAÇÃO] Lendo produtos de {cod}...")

        while True:
            if automation_status['request_stop']:
                break
            try:
                BASE = f'/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[2]/div[2]/table/tbody[2]/tr/td/table/tbody/tr[{linha_idx}]'

                # Nome do produto (td[3])
                nome_el = self.driver.find_elements(By.XPATH, BASE + '/td[3]/a')
                if not nome_el or not nome_el[0].text:
                    break
                nome_produto = nome_el[0].text

                # Número REAL do produto (td[2] — link com o número visível)
                num_el = self.driver.find_elements(By.XPATH, BASE + '/td[2]/a')
                if not num_el:
                    break
                num_real = num_el[0].text.strip()
                try:
                    num_real = int(num_real)
                except ValueError:
                    num_real = linha_idx  # fallback

                num_el[0].click()
                self._aguardar_detalhe_produto(timeout=5)

                cielo_sit     = self._get_status(XP_SITUACAO_CIELO)  or '—'
                cielo_hab     = self._get_status(XP_STATUS_CIELO)    or '—'
                cardse_sit    = self._get_status(XP_SITUACAO_CARDSE) or '—'
                cardse_hab    = self._get_status(XP_STATUS_CARDSE)   or '—'

                self.add_log(f"  [#{num_real}] {nome_produto} | CIELO: sit={cielo_sit} hab={cielo_hab} | CARDSE: sit={cardse_sit} hab={cardse_hab}")
                resultado.append({'num': num_real, 'nome': nome_produto,
                                  'cielo_sit': cielo_sit, 'cielo_hab': cielo_hab,
                                  'cardse_sit': cardse_sit, 'cardse_hab': cardse_hab,
                                  'obs': ''})

                self._voltar_para_lista_produtos()
                linha_idx += 1

            except (NoSuchElementException, TimeoutException):
                break
            except Exception as e:
                if _is_session_error(e): raise
                self.handle_exception(e, f"Erro verificação linha #{linha_idx}")
                time.sleep(2); linha_idx += 1

        self.add_log(f"[VERIFICAÇÃO] {len(resultado)} produtos lidos para {cod}.")
        return resultado

    # ── Modo ATIVAÇÃO ─────────────────────────────────────────

    def _processar_produtos(self, cod, tecnologia, acao_cielo, acao_cardse):
        linha_idx       = 1   # índice da linha na tabela HTML
        total_alterados = 0
        nome_produto    = 'desconhecido'
        self.add_log("Processando produtos...")

        if self.produto_inicio > 1:
            self.add_log(f"Pulando produtos com número < {self.produto_inicio} conforme configuração.")

        while True:
            if automation_status['request_stop']:
                break
            try:
                BASE = f'/html/body/div[2]/div/form/div/div[1]/section/div[2]/section[2]/div[2]/table/tbody[2]/tr/td/table/tbody/tr[{linha_idx}]'

                # Nome do produto (td[3])
                nome_el = self.driver.find_elements(By.XPATH, BASE + '/td[3]/a')
                if not nome_el or not nome_el[0].text:
                    break
                nome_produto = nome_el[0].text

                # Número REAL do produto (td[2])
                num_el = self.driver.find_elements(By.XPATH, BASE + '/td[2]/a')
                if not num_el:
                    break
                try:
                    num_real = int(num_el[0].text.strip())
                except ValueError:
                    num_real = linha_idx  # fallback

                # Pula produtos cujo número real é menor que produto_inicio
                if num_real < int(self.produto_inicio):
                    self.add_log(f"  [SKIP] Produto [#{num_real}]: {nome_produto} (início configurado: #{self.produto_inicio})")
                    linha_idx += 1
                    continue

                self.add_log(f"Produto [#{num_real}]: {nome_produto}")
                num_el[0].click()
                self._aguardar_detalhe_produto(timeout=5)

                self._aprovar_se_necessario()
                alterado = self._executar_acao_tecnologia(nome_produto, num_real, tecnologia, acao_cielo, acao_cardse)
                if alterado:
                    total_alterados += 1
                    self.add_log(f"✓ Produto [#{num_real}] '{nome_produto}' processado!", level="SUCCESS")
                linha_idx += 1

            except (NoSuchElementException, TimeoutException):
                break
            except Exception as e:
                if _is_session_error(e):
                    self.add_log("Sessão do navegador perdida durante processamento de produtos.", level="ERROR")
                    raise
                self.add_log(
                    f"ATENÇÃO: Erro no produto [#{linha_idx}] '{nome_produto}' — "
                    f"total alterado até agora: [{total_alterados}]", level="ERROR")
                self.handle_exception(e, f"Erro produto linha #{linha_idx}")
                time.sleep(2); linha_idx += 1

        self.add_log(f"Produtos: {total_alterados} alterados no total.")

    # ── Execução de ações ─────────────────────────────────────

    def _executar_acao_tecnologia(self, nome, contador, tecnologia, acao_cielo, acao_cardse):
        try:
            if   tecnologia == 1: return self._executar_cielo(nome, contador, acao_cielo)
            elif tecnologia == 2: return self._executar_cardse(nome, contador, acao_cardse)
            else:                 return self._executar_ambos(nome, contador, acao_cielo, acao_cardse)
        except Exception as e:
            self.handle_exception(e, f"Erro ação #{contador}"); return False

    def _confirmar_popup_desfazer(self):
        """Verifica se apareceu o pop-up de confirmação ao desfazer/desabilitar e clica em SIM."""
        # Estratégias em ordem de prioridade — ID dinâmico, texto, posição
        XPATHS_SIM = [
            # Pelo texto "Sim" dentro de qualquer botão do messagebox
            '//div[contains(@id,"messagebox")]//span[normalize-space(text())="Sim"]',
            '//div[contains(@id,"messagebox")]//button[normalize-space(text())="Sim"]',
            # Pelo ID fixo informado (tenta mesmo assim)
            '//*[@id="button-1006-btnIconEl"]',
            '//*[@id="button-1006-btnInnerEl"]',
            '//*[@id="button-1006"]',
            # Primeiro botão visível do messagebox (SIM costuma ser o primeiro)
            '//div[contains(@id,"messagebox")]//a[contains(@class,"x-btn")][1]',
            '//div[contains(@id,"messagebox")]//span[contains(@class,"x-btn-inner")][1]',
        ]
        try:
            # Primeiro aguarda o messagebox aparecer
            WebDriverWait(self.driver, 3).until(
                EC.presence_of_element_located((By.XPATH, '//div[contains(@id,"messagebox")]'))
            )
            self.add_log("Pop-up de confirmação detectado → clicando SIM...")

            clicou = False
            for xp in XPATHS_SIM:
                try:
                    el = WebDriverWait(self.driver, 1).until(
                        EC.presence_of_element_located((By.XPATH, xp))
                    )
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", el)
                    self.driver.execute_script("arguments[0].click();", el)
                    time.sleep(0.4)
                    self.add_log("Confirmação SIM aceita.")
                    clicou = True
                    break
                except:
                    continue

            if not clicou:
                # Última tentativa: pressionar Enter (foca no botão padrão do dialog)
                self.add_log("Tentando confirmar via teclado (Enter)...")
                from selenium.webdriver.common.keys import Keys
                webdriver.ActionChains(self.driver).send_keys(Keys.RETURN).perform()
                time.sleep(0.5)

        except (TimeoutException, NoSuchElementException):
            pass  # Pop-up não apareceu, segue normalmente

    def _confirmar_popup_ativacao_cielo(self):
        """Confirma o pop-up especÃ­fico que pode abrir ao ativar na Cielo."""
        try:
            botao = WebDriverWait(self.driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, XP_POPUP_CIELO_OK))
            )
            self.add_log("Pop-up da CIELO detectado apÃ³s ativar â†’ confirmando...")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
            self.driver.execute_script("arguments[0].click();", botao)
            time.sleep(1.5)
            self.add_log("Pop-up da CIELO confirmado.")
        except (TimeoutException, NoSuchElementException):
            pass

    def _ativar_cielo(self):
        self.driver.find_element(By.XPATH, XP_BTN_CIELO).click()
        time.sleep(2)
        self._confirmar_popup_ativacao_cielo()

    def _confirmar_popup_ativacao_cielo_rapido(self):
        """VersÃ£o mais rÃ¡pida da confirmaÃ§Ã£o do pop-up de ativaÃ§Ã£o da Cielo."""
        fim = time.time() + 1.2
        while time.time() < fim:
            try:
                botoes = self.driver.find_elements(By.XPATH, XP_POPUP_CIELO_OK)
                for botao in botoes:
                    if botao.is_displayed() and botao.is_enabled():
                        self.add_log("Pop-up da CIELO detectado apÃ³s ativar â†’ confirmando...")
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
                        self.driver.execute_script("arguments[0].click();", botao)
                        time.sleep(0.4)
                        self.add_log("Pop-up da CIELO confirmado.")
                        return
            except Exception:
                pass
            time.sleep(0.15)

    def _ativar_cielo_rapido(self):
        self.driver.find_element(By.XPATH, XP_BTN_CIELO).click()
        time.sleep(0.5)
        self._confirmar_popup_ativacao_cielo_rapido()

    def _confirmar_popup_ativacao_cielo_rapido(self):
        xpaths_ok = [
            XP_POPUP_CIELO_OK,
            '//*[@id="button-1006-btnInnerEl"]',
            '//*[@id="button-1006"]',
            '//div[contains(@id,"messagebox")]//span[normalize-space(text())="OK"]',
            '//div[contains(@id,"messagebox")]//button[normalize-space(text())="OK"]',
            '//div[contains(@id,"messagebox")]//a[contains(@class,"x-btn")][1]',
            '//div[contains(@id,"messagebox")]//span[contains(@class,"x-btn-inner")][1]',
        ]
        fim = time.time() + 2.0
        implicit_original = getattr(self, '_implicit_wait', 0)
        try:
            self.driver.implicitly_wait(0)
            while time.time() < fim:
                for xp in xpaths_ok:
                    try:
                        botoes = self.driver.find_elements(By.XPATH, xp)
                        for botao in botoes:
                            if botao.is_displayed() and botao.is_enabled():
                                self.add_log("Pop-up da CIELO detectado apos ativar; confirmando...")
                                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
                                self.driver.execute_script("arguments[0].click();", botao)
                                self.add_log("Pop-up da CIELO confirmado.")
                                return True
                    except Exception:
                        pass
                time.sleep(0.15)
        finally:
            self.driver.implicitly_wait(implicit_original)
        return False

    def _ativar_cielo_rapido(self):
        self.driver.find_element(By.XPATH, XP_BTN_CIELO).click()
        self._confirmar_popup_ativacao_cielo_rapido()
        if not self._aguardar_status(XP_STATUS_CIELO, 'SIM', timeout=4):
            self.add_log("Aviso: CIELO nao confirmou status SIM dentro do tempo esperado.", level="WARNING")

    def _executar_cielo(self, nome, contador, acao):
        s = self._get_status(XP_STATUS_CIELO)
        habilitado = s.upper() == 'SIM'
        self.add_log(f"CIELO [{contador}]: '{nome}' | Habilitado: {s or 'vazio'} | Ação: {acao}")
        if acao == 'ATIVAR':
            if habilitado:
                self.add_log("CIELO já habilitado — sem ação.")
                self._voltar_para_lista_produtos(); return False
            self.add_log("CIELO desabilitado → clicando HABILITAR")
            self._ativar_cielo_rapido()
            self._voltar_para_lista_produtos(); return True
        else:
            if not habilitado:
                self.add_log("CIELO já desabilitado — sem ação.")
                self._voltar_para_lista_produtos(); return False
            self.add_log("CIELO habilitado → clicando DESABILITAR")
            self.driver.find_element(By.XPATH, XP_BTN_CIELO).click()
            self._confirmar_popup_desfazer()
            self._aguardar_status(XP_STATUS_CIELO, 'NAO', timeout=4)
            self._voltar_para_lista_produtos(); return True

    def _desabilitar_cardse_com_fallback_ativado(self):
        """Clica no botão DESABILITAR do CARDSE, confirma o popup SIM e, se após
        isso o status ainda for SIM (situação ENVIADO → ATIVADO), verifica se ATIVADO
        também está na lista de situações permitidas e clica novamente para finalizar."""
    def _desabilitar_cardse_com_fallback_ativado(self):
        """Clica no botão DESABILITAR do CARDSE, confirma o popup SIM e, se após
        isso o status ainda for SIM (situação ENVIADO → ATIVADO), verifica se ATIVADO
        também está na lista de situações permitidas e clica novamente para finalizar."""
        self.driver.find_element(By.XPATH, XP_BTN_CARDSE).click()
        self._confirmar_popup_desfazer()
        # Não aguarda tempo fixo; o próximo passo verifica o status atual.
        # Se ainda estiver habilitado, tenta desativar novamente conforme regra.
        status_apos = (self._get_status(XP_STATUS_CARDSE) or '').upper()
        if status_apos == 'SIM':
            situacao_apos = self._get_status(XP_SITUACAO_CARDSE)
            self.add_log(f"CARDSE ainda habilitado após SIM (situação atual: {situacao_apos}) — verificando se deve desabilitar novamente...")
            if self._cardse_situacao_permitida(situacao_apos):
                self.add_log(f"Situação '{situacao_apos}' também permitida para DESATIVAR → clicando DESABILITAR novamente.")
                self.driver.find_element(By.XPATH, XP_BTN_CARDSE).click()
                self._confirmar_popup_desfazer()
                self._aguardar_status(XP_STATUS_CARDSE, 'NAO', timeout=4)
            else:
                self.add_log(f"Situação '{situacao_apos}' NÃO está na lista para DESATIVAR — encerrando sem nova ação.")
        else:
            self._aguardar_status(XP_STATUS_CARDSE, 'NAO', timeout=4)

    def _executar_cardse(self, nome, contador, acao):
        s = self._get_status(XP_STATUS_CARDSE)
        habilitado = s.upper() == 'SIM'
        self.add_log(f"CARDSE [{contador}]: '{nome}' | Habilitado: {s or 'vazio'} | Ação: {acao}")
        if acao == 'ATIVAR':
            if habilitado:
                self.add_log("CARDSE já habilitado — sem ação.")
                self._voltar_para_lista_produtos(); return False
            self.add_log("CARDSE desabilitado → clicando HABILITAR")
            self._clicar_rede_captura(XP_BTN_CARDSE, XP_STATUS_CARDSE, 'SIM', 'CARDSE', timeout=4)
            self._voltar_para_lista_produtos(); return True
        else:
            if not habilitado:
                self.add_log("CARDSE já desabilitado — sem ação.")
                self._voltar_para_lista_produtos(); return False
            if not self._cardse_situacao_permitida(self._get_status(XP_SITUACAO_CARDSE)):
                self.add_log(f"CARDSE habilitado, mas situação não permitida para DESATIVAR ({self._get_status(XP_SITUACAO_CARDSE)}) — pulando.")
                self._voltar_para_lista_produtos(); return False
            self.add_log("CARDSE habilitado → clicando DESABILITAR")
            self._desabilitar_cardse_com_fallback_ativado()
            self._voltar_para_lista_produtos(); return True

    def _executar_ambos(self, nome, contador, acao_cielo, acao_cardse):
        alterou = False
        sc = self._get_status(XP_STATUS_CIELO)
        hab_c = sc.upper() == 'SIM'
        self.add_log(f"CIELO [{contador}]: '{nome}' | Habilitado: {sc or 'vazio'} | Ação: {acao_cielo}")
        if acao_cielo == 'ATIVAR':
            if hab_c: self.add_log("CIELO já habilitado — sem ação.")
            else:
                self.add_log("CIELO desabilitado → HABILITAR")
                self._ativar_cielo_rapido(); alterou = True
        else:
            if not hab_c: self.add_log("CIELO já desabilitado — sem ação.")
            else:
                self.add_log("CIELO habilitado → DESABILITAR")
                self.driver.find_element(By.XPATH, XP_BTN_CIELO).click()
                self._confirmar_popup_desfazer(); alterou = True
                self._aguardar_status(XP_STATUS_CIELO, 'NAO', timeout=4)

        ss = self._get_status(XP_STATUS_CARDSE)
        hab_s = ss.upper() == 'SIM'
        self.add_log(f"CARDSE [{contador}]: '{nome}' | Habilitado: {ss or 'vazio'} | Ação: {acao_cardse}")
        if acao_cardse == 'ATIVAR':
            if hab_s: self.add_log("CARDSE já habilitado — sem ação.")
            else:
                self.add_log("CARDSE desabilitado → HABILITAR")
                self._clicar_rede_captura(XP_BTN_CARDSE, XP_STATUS_CARDSE, 'SIM', 'CARDSE', timeout=4); alterou = True
        else:
            if not hab_s:
                self.add_log("CARDSE já desabilitado — sem ação.")
            elif not self._cardse_situacao_permitida(self._get_status(XP_SITUACAO_CARDSE)):
                self.add_log(f"CARDSE habilitado, mas situação não permitida para DESATIVAR ({self._get_status(XP_SITUACAO_CARDSE)}) — pulando.")
            else:
                self.add_log("CARDSE habilitado → DESABILITAR")
                self._desabilitar_cardse_com_fallback_ativado(); alterou = True

        self._voltar_para_lista_produtos()
        return alterou

    # ── Finalização ───────────────────────────────────────────

    def _finalizar(self):
        # Sempre fecha o driver, mesmo se houver erro
        if self.driver:
            try:
                self.driver.quit()
                self.add_log("Navegador fechado.")
            except Exception as e:
                self.handle_exception(e, "Erro ao fechar navegador")

        # Gera o relatório apropriado
        if self.modo_coleta_estabelecimento:
            if self._relatorio_path:
                try:
                    self._salvar_relatorio_coleta(parcial=False, silencioso=True)
                    self.add_log(f"Planilha gerada: {os.path.basename(self._relatorio_path)}", level="SUCCESS")
                except Exception as e:
                    self.handle_exception(e, "Erro ao gerar planilha da coleta")
        elif self.modo_verificacao and self._dados_relatorio:
            ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
            caminho = os.path.join(os.getcwd(), f"verificacao_{ts}.xlsx")
            try:
                _gerar_relatorio_xlsx(self._dados_relatorio, caminho)
                automation_status['relatorio_path'] = caminho
                self.add_log(f"Planilha gerada: {os.path.basename(caminho)}", level="SUCCESS")
            except Exception as e:
                self.handle_exception(e, "Erro ao gerar planilha")

        # Log final
        tempo = "N/A"
        if automation_status['start_time'] and automation_status['end_time']:
            tempo = str(automation_status['end_time'] - automation_status['start_time']).split('.')[0]
        total_final = automation_status['total'] or 1  # evita divisão por zero
        processed_final = automation_status['processed']
        self.add_log(f"Concluído. Processados: {processed_final}/{total_final} | Tempo: {tempo}")
        automation_status['request_stop'] = False
